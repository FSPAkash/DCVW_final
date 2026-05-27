"""
Fulfillment ledger for DCW SIOP v3.

Each fulfillment = one commit. Persisted to fulfillments.json (newest first).
Within 24h of original commit, a fulfillment is editable: lines can be
added/removed, swapped to a different lot, or have consume_mt changed. Edits
atomically rewind the previous deductions on the lots involved and re-apply
the new ones to Master.xlsx.
"""
from __future__ import annotations

import json
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

import v3_master

BASE_DIR = Path(__file__).resolve().parent
LEDGER_FILE = BASE_DIR / "fulfillments.json"
EDIT_WINDOW_HOURS = 24


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _load() -> List[Dict[str, Any]]:
    if not LEDGER_FILE.exists():
        return []
    try:
        return json.loads(LEDGER_FILE.read_text(encoding="utf-8"))
    except Exception:
        return []


def _save(rows: List[Dict[str, Any]]) -> None:
    LEDGER_FILE.write_text(json.dumps(rows, indent=2, default=str), encoding="utf-8")


def is_editable(fulfillment: Dict[str, Any]) -> bool:
    try:
        ts = datetime.fromisoformat(fulfillment["ts"])
    except Exception:
        return False
    return datetime.now() - ts < timedelta(hours=EDIT_WINDOW_HOURS)


def list_all() -> List[Dict[str, Any]]:
    rows = _load()
    # Newest first
    rows.sort(key=lambda r: r.get("ts", ""), reverse=True)
    for r in rows:
        r["editable"] = is_editable(r)
    return rows


def record_commit(*, user: str, customer_id: str, customer_name: str, qty_requested: float,
                   write_result: Dict[str, Any],
                   snapshot_by_lot: Optional[Dict[str, Dict[str, Any]]] = None) -> Dict[str, Any]:
    """Called after v3_master.update_lot_qty succeeds. Persists the line items.

    snapshot_by_lot: optional {lot_id: {ranks, mass_tone, tint_tone, within,
    direction, is_super, present_methods, ...}} captured from the /match result
    that was on screen at commit time. Stored on each line so that edit-mode
    can show the original ranking/spec/dir info even after the lot is depleted.
    """
    snap = snapshot_by_lot or {}
    rows = _load()
    entry = {
        "id": uuid.uuid4().hex[:12],
        "ts": _now(),
        "user": user,
        "customer_id": customer_id,
        "customer_name": customer_name,
        "qty_requested": qty_requested,
        "lines": [{
            "lot_id": a["lot_id"],
            "lot_no": a["lot_no"],
            "col_letter": a.get("col_letter"),
            "consume_mt": a["consume_mt"],
            "prev_qty": a["prev_qty"],
            "new_qty": a["new_qty"],
            "snapshot": snap.get(a["lot_id"]),
        } for a in write_result.get("applied", [])],
        "edit_history": [],
    }
    rows.append(entry)
    _save(rows)
    return entry


# ----- Editing -----

def _find(fid: str) -> Tuple[Optional[Dict[str, Any]], List[Dict[str, Any]]]:
    rows = _load()
    for r in rows:
        if r["id"] == fid:
            return r, rows
    return None, rows


def _adjust_lot_qty(lot_id_or_no: str, delta: float, ws, id_to_col: Dict[str, int],
                     name_to_cols: Dict[str, List[int]]) -> Tuple[float, float, int]:
    """Add `delta` to a lot's QTY cell + stamp 'Last edited' row. Returns (prev, new, col)."""
    col = id_to_col.get(lot_id_or_no)
    if col is None:
        cols = name_to_cols.get(lot_id_or_no, [])
        if len(cols) == 1:
            col = cols[0]
    if col is None:
        raise ValueError(f"lot not found: {lot_id_or_no}")
    cell = ws.cell(v3_master.ROW_QTY, col)
    prev = float(cell.value) if isinstance(cell.value, (int, float)) else (v3_master._to_num(cell.value) or 0.0)
    new = round(prev + delta, 3)
    if new < 0:
        new = 0.0
    cell.value = new
    ws.cell(v3_master.ROW_LAST_EDITED, col).value = datetime.now().strftime("%Y-%m-%d")
    return prev, new, col


def _build_col_maps(ws) -> Tuple[Dict[str, int], Dict[str, List[int]]]:
    id_to_col: Dict[str, int] = {}
    name_to_cols: Dict[str, List[int]] = {}
    for c in range(1, ws.max_column + 1):
        v = v3_master._norm(ws.cell(v3_master.ROW_LOT, c).value)
        if v and v.upper() != "LOT":
            id_to_col[f"{v}@{v3_master._col_letter(c)}"] = c
            name_to_cols.setdefault(v, []).append(c)
    return id_to_col, name_to_cols


def append_new_lot(ws, *, lot_no: str, grade: str, qty_mt: float,
                    mt: Dict[str, float], rt: Dict[str, float]) -> Dict[str, Any]:
    """Append a new lot column to the 'proposed' sheet with MT (Method I a) + RT (Method I b) values.
    Returns {col, col_letter, lot_id}."""
    new_col = ws.max_column + 1
    # Headers
    ws.cell(v3_master.ROW_GRADE, new_col).value = grade
    ws.cell(v3_master.ROW_LOT, new_col).value = lot_no
    ws.cell(v3_master.ROW_QTY, new_col).value = float(qty_mt)
    # Method I a (rows 9-12: DL, Da, Db, DE)
    for r, k in [(9, "DL"), (10, "Da"), (11, "Db"), (12, "DE")]:
        v = mt.get(k)
        ws.cell(r, new_col).value = float(v) if v is not None else None
    # Method I b (rows 13-17: DL, Da, Db, DE, Strength)
    for r, k in [(13, "DL"), (14, "Da"), (15, "Db"), (16, "DE"), (17, "Strength")]:
        v = rt.get(k)
        ws.cell(r, new_col).value = float(v) if v is not None else None
    ws.cell(v3_master.ROW_LAST_EDITED, new_col).value = datetime.now().strftime("%Y-%m-%d")
    return {
        "col": new_col,
        "col_letter": v3_master._col_letter(new_col),
        "lot_id": f"{lot_no}@{v3_master._col_letter(new_col)}",
        "lot_no": lot_no,
    }


def edit_fulfillment(fid: str, *, user: str, new_lines: List[Dict[str, Any]],
                      new_lots: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Any]:
    """Atomically:
       1) Rewind every original line (add consume_mt back to original lot).
       2) For each new_lot in new_lots, append column to Master.
       3) Apply each new_line (deduct consume_mt from its lot).
       Replace the fulfillment's lines with new_lines.

       Returns updated fulfillment record.
    """
    entry, rows = _find(fid)
    if entry is None:
        raise ValueError(f"fulfillment not found: {fid}")
    if not is_editable(entry):
        raise PermissionError("fulfillment is past 24h edit window")

    wb = v3_master._load_wb(read_only=False)
    ws = wb[v3_master.SHEET_NAME]
    # Ensure "Last edited" label exists
    label_cell = ws.cell(v3_master.ROW_LAST_EDITED, v3_master.LAST_EDITED_LABEL_COL)
    if v3_master._norm(label_cell.value).lower() != v3_master.LAST_EDITED_LABEL.lower():
        label_cell.value = v3_master.LAST_EDITED_LABEL

    # Two-pass approach: (1) fully rewind every previous line, (2) append any
    # new lots, (3) apply every new line. This makes the ledger's prev_qty /
    # new_qty / consume_mt internally consistent (prev - consume == new) and
    # decouples bookkeeping from net-delta arithmetic.
    id_to_col, name_to_cols = _build_col_maps(ws)

    prev_lines = entry.get("lines", []) or []
    prev_by_key: Dict[str, float] = {}
    for line in prev_lines:
        k = line.get("lot_id") or line.get("lot_no")
        if not k:
            continue
        prev_by_key[k] = prev_by_key.get(k, 0.0) + float(line.get("consume_mt") or 0)

    # PASS 1: rewind every previous line.
    rewinds: List[Dict[str, Any]] = []
    for key, prev_consume in prev_by_key.items():
        if prev_consume <= 0:
            continue
        try:
            _, after_rewind, col = _adjust_lot_qty(key, +prev_consume, ws, id_to_col, name_to_cols)
            real_lot_no = v3_master._norm(ws.cell(v3_master.ROW_LOT, col).value)
            real_col_letter = v3_master._col_letter(col)
            rewinds.append({
                "lot_id": f"{real_lot_no}@{real_col_letter}",
                "restored": prev_consume,
                "qty_now": after_rewind,
            })
        except ValueError as e:
            rewinds.append({"lot_id": key, "restored": 0, "qty_now": None, "error": str(e)})

    # PASS 2: append any brand-new lots (new column on the sheet).
    appended: List[Dict[str, Any]] = []
    for nl in (new_lots or []):
        info = append_new_lot(ws, lot_no=nl["lot_no"], grade=nl.get("grade", ""),
                              qty_mt=float(nl.get("qty_mt", 0)),
                              mt=nl.get("mass_tone") or {},
                              rt=nl.get("tint_tone") or {})
        appended.append(info)
    id_to_col, name_to_cols = _build_col_maps(ws)

    # PASS 3: apply every new line against the now-rewound master.
    new_by_key: Dict[str, float] = {}
    for line in new_lines:
        k = line.get("lot_id") or line.get("lot_no")
        if not k:
            continue
        new_by_key[k] = new_by_key.get(k, 0.0) + float(line.get("consume_mt") or 0)

    applied: List[Dict[str, Any]] = []
    for key, new_consume in new_by_key.items():
        if new_consume <= 0:
            continue
        try:
            prev_qty, after_apply, col = _adjust_lot_qty(key, -new_consume, ws, id_to_col, name_to_cols)
            real_lot_no = v3_master._norm(ws.cell(v3_master.ROW_LOT, col).value)
            real_col_letter = v3_master._col_letter(col)
            applied.append({
                "lot_id": f"{real_lot_no}@{real_col_letter}",
                "lot_no": real_lot_no,
                "col_letter": real_col_letter,
                "consume_mt": new_consume,
                "prev_qty": prev_qty,
                "new_qty": after_apply,
            })
        except ValueError as e:
            applied.append({"lot_id": key, "error": str(e)})

    wb.save(v3_master.MASTER_FILE)

    # Stamp edit_dates for everything touched
    edit_dates = v3_master._read_edit_dates()
    today = datetime.now().strftime("%Y-%m-%d")
    for r in rewinds + applied:
        if r.get("lot_id"):
            edit_dates[r["lot_id"]] = today
    v3_master._write_edit_dates(edit_dates)

    # Update ledger entry
    entry.setdefault("edit_history", []).append({
        "ts": _now(),
        "user": user,
        "rewinds": rewinds,
        "appended_lots": appended,
        "applied": applied,
        "previous_lines": list(entry.get("lines", [])),
    })
    # Preserve original snapshots so depleted lots keep their rank/spec info
    # through the 24h edit window across multiple edits.
    prev_snap_by_id = {pl.get("lot_id"): pl.get("snapshot")
                       for pl in (entry.get("lines") or [])}
    entry["lines"] = [{
        "lot_id": a["lot_id"],
        "lot_no": a.get("lot_no"),
        "col_letter": a.get("col_letter"),
        "consume_mt": a.get("consume_mt"),
        "prev_qty": a.get("prev_qty"),
        "new_qty": a.get("new_qty"),
        "snapshot": prev_snap_by_id.get(a["lot_id"]),
    } for a in applied if "error" not in a]
    entry["edited_at"] = _now()
    _save(rows)
    return entry


if __name__ == "__main__":
    for r in list_all()[:5]:
        print(r["ts"], r["customer_name"], "lines:", len(r["lines"]), "editable:", r["editable"])
