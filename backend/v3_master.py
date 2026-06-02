"""
Master.xlsx parser + writer for DCW SIOP v3.

Master file structure (sheet 'proposed '):
  Row 6: Grade per lot (cols C..)
  Row 7: LOT no per lot (col B='LOT', cols C.. = lot codes)
  Row 8: QTY(MT) per lot (col C='QTY(MT)')
  Rows 9..55: method blocks. Each block has:
    col B (only on first row of block) = Method label e.g. 'Method I a'
    col C = submethod family e.g. 'Alk MT '
    col D = axis (DL/Da/Db/DE/Strength)
    cols E.. = values per lot

Each lot is a column. We pivot to per-lot records.

We also keep a 'last_edited' row (added if missing) under QTY.
"""
from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
import runtime_paths

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = runtime_paths.ensure_data_layout()
MASTER_FILE = DATA_DIR / "Master.xlsx"
AUDIT_FILE = DATA_DIR / "inventory_audit.csv"
SHEET_NAME = "proposed "

# Row indices (1-based)
ROW_GRADE = 6
ROW_LOT = 7
ROW_QTY = 8
ROW_METHODS_START = 9
ROW_METHODS_END = 55
# Rows 56-60 are 'Special Analysis' labels. We use row 61 for inventory edit dates.
ROW_LAST_EDITED = 61
LAST_EDITED_LABEL_COL = 4  # col D, matches axis-label column
LAST_EDITED_LABEL = "Last edited"

EDIT_DATES_FILE = DATA_DIR / "inventory_edit_dates.json"
BLEND_RECORDS_FILE = DATA_DIR / "blend_records.json"
BLEND_CARDS_DIR = DATA_DIR / "blend_cards"
BLEND_CARDS_DIR.mkdir(parents=True, exist_ok=True)


def _norm(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()


def _to_num(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.upper() in {"N/A", "NA", "-", "XX"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _load_wb(read_only: bool = True):
    return openpyxl.load_workbook(MASTER_FILE, data_only=read_only)


def _read_edit_dates() -> Dict[str, str]:
    if EDIT_DATES_FILE.exists():
        try:
            return json.loads(EDIT_DATES_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def _write_edit_dates(d: Dict[str, str]) -> None:
    EDIT_DATES_FILE.write_text(json.dumps(d, indent=2), encoding="utf-8")


def _read_blend_records() -> Dict[str, Any]:
    if BLEND_RECORDS_FILE.exists():
        try:
            data = json.loads(BLEND_RECORDS_FILE.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                data.setdefault("blends", [])
                data.setdefault("notes_by_lot", {})
                return data
        except Exception:
            pass
    return {"blends": [], "notes_by_lot": {}}


def _write_blend_records(d: Dict[str, Any]) -> None:
    BLEND_RECORDS_FILE.write_text(json.dumps(d, indent=2), encoding="utf-8")


def read_blend_records() -> Dict[str, Any]:
    return _read_blend_records()


def _col_letter(col: int) -> str:
    """1-based col index -> Excel letter (A, B, ..., AA, ...)."""
    s = ""
    while col > 0:
        col, rem = divmod(col - 1, 26)
        s = chr(65 + rem) + s
    return s


def parse_master() -> Dict[str, Any]:
    """Return {'lots': [...], 'methods': [...], 'submethods_by_method': {...}, 'duplicates': [...]}."""
    wb = _load_wb(read_only=True)
    ws = wb[SHEET_NAME]
    max_col = ws.max_column

    # Build method-row layout
    # For each row in methods range: record (method_label_or_inherit, submethod, axis)
    layout: List[Dict[str, str]] = []
    current_method = None
    for r in range(ROW_METHODS_START, ROW_METHODS_END + 1):
        m = _norm(ws.cell(r, 2).value)
        sub = _norm(ws.cell(r, 3).value)
        axis = _norm(ws.cell(r, 4).value)
        if m:
            current_method = m
        layout.append({"row": r, "method": current_method, "submethod": sub, "axis": axis})

    # Collect distinct methods preserving order
    methods: List[str] = []
    submethod_of_method: Dict[str, str] = {}
    for row in layout:
        if row["method"] and row["method"] not in methods:
            methods.append(row["method"])
            submethod_of_method[row["method"]] = row["submethod"]

    edit_dates = _read_edit_dates()
    blend_records = _read_blend_records()
    notes_by_lot = blend_records.get("notes_by_lot", {})
    blend_lot_ids = {b.get("output", {}).get("lot_id") for b in blend_records.get("blends", [])}

    # Iterate lot columns (col 5 onwards; col 2-4 are labels, col 1 blank)
    # Actually data starts col 3 ('LOT' header is col 2='LOT'? let's recheck)
    # From inspection: R7 cells: B7=None? Let's check: B='LOT' at C7? earlier dump said R7 C='LOT'? no.
    # The earlier dump for proposed printed row 7 starting at A. cols: A,B blank, then ['LOT','6210B'...]
    # Lookup: in our row-dumper we did `r[0], r[1]` for A,B. That printed both None. Then values_only row was
    # ['LOT','6210B',...] starting where? Below code uses cells; safer: scan row 7 across cols.
    lot_col_start = None
    for c in range(1, max_col + 1):
        v = _norm(ws.cell(ROW_LOT, c).value)
        if v and v.upper() == "LOT":
            lot_col_start = c + 1
            break
    if lot_col_start is None:
        # fallback to col 5
        lot_col_start = 5

    lots: List[Dict[str, Any]] = []
    for c in range(lot_col_start, max_col + 1):
        lot_no = _norm(ws.cell(ROW_LOT, c).value)
        if not lot_no:
            continue
        grade = _norm(ws.cell(ROW_GRADE, c).value)
        qty = _to_num(ws.cell(ROW_QTY, c).value)
        if qty is None:
            qty = 0.0

        # Group axis values by method
        blocks: Dict[str, Dict[str, Optional[float]]] = {}
        for row in layout:
            method = row["method"] or "Unknown"
            axis = row["axis"]
            if not axis:
                continue
            val = _to_num(ws.cell(row["row"], c).value)
            blocks.setdefault(method, {})[axis] = val

        # Determine which methods are "present" (have at least one numeric value)
        present_methods = [m for m, axes in blocks.items() if any(v is not None for v in axes.values())]

        cell_edit = _norm(ws.cell(ROW_LAST_EDITED, c).value)
        if isinstance(ws.cell(ROW_LAST_EDITED, c).value, datetime):
            cell_edit = ws.cell(ROW_LAST_EDITED, c).value.strftime("%Y-%m-%d")
        lid = f"{lot_no}@{_col_letter(c)}"
        lots.append({
            "col": c,
            "col_letter": _col_letter(c),
            "lot_id": lid,  # globally unique key
            "lot_no": lot_no,
            "grade": grade,
            "qty_mt": qty,
            "blocks": blocks,
            "present_methods": present_methods,
            "last_edited": cell_edit or edit_dates.get(lid) or edit_dates.get(lot_no),
            "blend_notes": notes_by_lot.get(lid, []),
            "is_blended": lid in blend_lot_ids,
        })

    # Detect duplicate lot codes (data integrity flag)
    from collections import Counter
    cnt = Counter(l["lot_no"] for l in lots)
    dup_codes = {k for k, v in cnt.items() if v > 1}
    duplicates = [
        {"lot_no": l["lot_no"], "col": l["col"], "col_letter": l["col_letter"], "grade": l["grade"], "qty_mt": l["qty_mt"]}
        for l in lots if l["lot_no"] in dup_codes
    ]

    # Blend cards must reflect what is actually in Master.xlsx, not the
    # append-only JSON log. A blend whose output lot was removed from Master
    # (e.g. deleted for testing) must not show a card. Anchor on live lots.
    live_lot_ids = {l["lot_id"] for l in lots}
    recorded_blend_ids = {
        b.get("output", {}).get("lot_id") for b in blend_records.get("blends", [])
    }

    # Backfill: a lot typed directly into Master with a BLEND prefix on its lot
    # code is a blend made outside the /master/blend endpoint, so it has no JSON
    # record yet. Synthesize a minimal record (no source breakdown is knowable
    # from the sheet) and persist it so it gets a card like any other blend.
    backfilled = False
    for l in lots:
        if l["lot_id"] in recorded_blend_ids:
            continue
        if not l["lot_no"].upper().startswith("BLEND"):
            continue
        rec_id = f"BLEND-DIRECT-{l['lot_id']}"
        record = {
            "blend_id": rec_id,
            "ts": l.get("last_edited") or "",
            "user": "master.xlsx",
            "direct_entry": True,
            "output": {
                "lot_id": l["lot_id"],
                "lot_no": l["lot_no"],
                "grade": l["grade"],
                "col_letter": l["col_letter"],
                "qty_mt": l["qty_mt"],
            },
            "sources": [],
            "card_url": f"/api/v3/master/blend/{rec_id}/card.pdf",
        }
        blend_records.setdefault("blends", []).append(record)
        recorded_blend_ids.add(l["lot_id"])
        backfilled = True
    if backfilled:
        _write_blend_records(blend_records)

    visible_blends = [
        b for b in blend_records.get("blends", [])
        if (b.get("output", {}).get("lot_id")) in live_lot_ids
    ]

    return {
        "lots": lots,
        "methods": methods,
        "submethod_of_method": submethod_of_method,
        "layout": layout,
        "lot_col_start": lot_col_start,
        "duplicates": duplicates,
        "blend_records": visible_blends,
    }


def update_lot_qty(updates: List[Dict[str, Any]], *, user: str = "system", customer: str = "") -> Dict[str, Any]:
    """
    updates: [{lot_no, consume_mt}, ...]
    Writes Master.xlsx in-place, updates edit_dates JSON, appends audit row.
    Returns summary {applied, skipped, after}.
    """
    wb = _load_wb(read_only=False)
    ws = wb[SHEET_NAME]
    max_col = ws.max_column

    # Build lot_id (lot_no@COL_LETTER) -> col AND lot_no -> [cols]
    id_to_col: Dict[str, int] = {}
    name_to_cols: Dict[str, List[int]] = {}
    for c in range(1, max_col + 1):
        v = _norm(ws.cell(ROW_LOT, c).value)
        if v and v.upper() != "LOT":
            id_to_col[f"{v}@{_col_letter(c)}"] = c
            name_to_cols.setdefault(v, []).append(c)

    edit_dates = _read_edit_dates()
    now_iso = datetime.now().isoformat(timespec="seconds")
    today = datetime.now().strftime("%Y-%m-%d")

    # Ensure "Last edited" label exists on row 61 (Excel column D)
    label_cell = ws.cell(ROW_LAST_EDITED, LAST_EDITED_LABEL_COL)
    if _norm(label_cell.value).lower() != LAST_EDITED_LABEL.lower():
        label_cell.value = LAST_EDITED_LABEL

    applied: List[Dict[str, Any]] = []
    skipped: List[Dict[str, Any]] = []

    audit_new = not AUDIT_FILE.exists()
    with AUDIT_FILE.open("a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if audit_new:
            w.writerow(["timestamp", "user", "customer", "lot_id", "lot_no", "col_letter", "prev_qty", "consume_mt", "new_qty"])

        for u in updates:
            lot_id = _norm(u.get("lot_id"))
            lot_no = _norm(u.get("lot_no"))
            consume = _to_num(u.get("consume_mt")) or 0.0
            if consume <= 0:
                skipped.append({"lot_id": lot_id or lot_no, "reason": "invalid_qty"})
                continue
            col = id_to_col.get(lot_id)
            if col is None and lot_no:
                cols = name_to_cols.get(lot_no, [])
                if len(cols) == 1:
                    col = cols[0]
                elif len(cols) > 1:
                    skipped.append({"lot_id": lot_id or lot_no, "reason": "ambiguous_lot_no"})
                    continue
            if col is None:
                skipped.append({"lot_id": lot_id or lot_no, "reason": "not_found"})
                continue
            prev = _to_num(ws.cell(ROW_QTY, col).value) or 0.0
            new_qty = round(prev - consume, 3)
            if new_qty < 0:
                new_qty = 0.0
            ws.cell(ROW_QTY, col).value = new_qty
            ws.cell(ROW_LAST_EDITED, col).value = today
            real_id = f"{lot_no or _norm(ws.cell(ROW_LOT, col).value)}@{_col_letter(col)}"
            edit_dates[real_id] = today
            w.writerow([now_iso, user, customer, real_id, lot_no or _norm(ws.cell(ROW_LOT, col).value), _col_letter(col), prev, consume, new_qty])
            applied.append({"lot_id": real_id, "lot_no": lot_no, "col_letter": _col_letter(col), "prev_qty": prev, "consume_mt": consume, "new_qty": new_qty})

    wb.save(MASTER_FILE)
    _write_edit_dates(edit_dates)

    return {"applied": applied, "skipped": skipped, "ts": now_iso}


def _blend_card_filename(blend_id: str) -> Path:
    safe = "".join(c for c in blend_id if c.isalnum() or c in ("-", "_"))
    return BLEND_CARDS_DIR / f"{safe}.pdf"


def generate_blend_card_pdf(record: Dict[str, Any]) -> Path:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    out = _blend_card_filename(record["blend_id"])
    doc = SimpleDocTemplate(str(out), pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm,
                            topMargin=18 * mm, bottomMargin=18 * mm, title=f"Blend {record['output']['lot_no']}")
    styles = getSampleStyleSheet()
    title = ParagraphStyle("title", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=18,
                           textColor=colors.HexColor("#8C2F1A"), spaceAfter=4)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=10.5,
                        textColor=colors.HexColor("#757575"), letterSpacing=1.5, spaceAfter=6, spaceBefore=12)
    body = ParagraphStyle("body", parent=styles["BodyText"], fontName="Helvetica", fontSize=11,
                          textColor=colors.HexColor("#111111"), spaceAfter=4)
    muted = ParagraphStyle("muted", parent=body, textColor=colors.HexColor("#757575"), fontSize=9.5)

    out_dict = record.get("output", {})
    sources = record.get("sources", [])

    story = []
    story.append(Paragraph(f"Blend · {out_dict.get('lot_no','')}", title))
    story.append(Paragraph(f"Grade {out_dict.get('grade','')}  ·  {out_dict.get('qty_mt',0)} MT  ·  column {out_dict.get('col_letter','')}", body))
    story.append(Paragraph(f"Created {record.get('ts','')} by {record.get('user','')}", muted))

    story.append(Paragraph("INVENTORY CHANGES", h2))
    rows = [["Lot", "Grade", "MT used"]]
    for s in sources:
        rows.append([f"{s.get('lot_no','')} @{s.get('lot_id','').split('@')[-1] if '@' in s.get('lot_id','') else ''}",
                     s.get("grade",""), f"{s.get('mt_used',0):.2f}"])
    rows.append(["", "Total used", f"{sum((s.get('mt_used') or 0) for s in sources):.2f}"])
    rows.append([f"{out_dict.get('lot_no','')} (new)", out_dict.get("grade",""), f"+{out_dict.get('qty_mt',0):.2f}"])
    t = Table(rows, colWidths=[70 * mm, 50 * mm, 35 * mm])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#757575")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#E5E5E5")),
        ("LINEBELOW", (0, 1), (-1, -3), 0.3, colors.HexColor("#EFEFEF")),
        ("LINEABOVE", (0, -1), (-1, -1), 0.8, colors.HexColor("#2E8B57")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F5EC")),
        ("TEXTCOLOR", (-1, -1), (-1, -1), colors.HexColor("#1F7A4A")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
    ]))
    story.append(t)

    story.append(Spacer(1, 6))
    story.append(Paragraph(
        "Master.xlsx was edited: source lot quantities reduced and a new column added for the blended lot.",
        muted,
    ))
    doc.build(story)
    return out


def get_blend_card_path(blend_id: str) -> Path:
    path = _blend_card_filename(blend_id)
    if path.exists():
        return path
    # rebuild if missing
    rec = _read_blend_records()
    for b in rec.get("blends", []):
        if b.get("blend_id") == blend_id:
            return generate_blend_card_pdf(b)
    raise FileNotFoundError(blend_id)


def compute_blend_preview(sources: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Beta auto weighted-average from source lots.
    sources: [{lot_id, mt_used}]
    Returns {"qty_mt": total, "methods": {method: {axis: weighted_value}}}.
    """
    data = parse_master()
    by_id = {l["lot_id"]: l for l in data["lots"]}
    weights: List[Dict[str, Any]] = []
    total_qty = 0.0
    for s in sources:
        lid = _norm(s.get("lot_id"))
        mt = _to_num(s.get("mt_used")) or 0.0
        lot = by_id.get(lid)
        if not lot or mt <= 0:
            continue
        weights.append({"lot": lot, "mt": mt})
        total_qty += mt
    methods_out: Dict[str, Dict[str, Optional[float]]] = {}
    if total_qty <= 0:
        return {"qty_mt": 0.0, "methods": methods_out}
    for method in data["methods"]:
        axis_map: Dict[str, Optional[float]] = {}
        axes_seen: set = set()
        for w in weights:
            axes_seen.update((w["lot"]["blocks"].get(method) or {}).keys())
        for axis in axes_seen:
            num = 0.0
            denom = 0.0
            for w in weights:
                v = (w["lot"]["blocks"].get(method) or {}).get(axis)
                if v is None:
                    continue
                num += v * w["mt"]
                denom += w["mt"]
            axis_map[axis] = round(num / denom, 4) if denom > 0 else None
        if any(v is not None for v in axis_map.values()):
            methods_out[method] = axis_map
    return {"qty_mt": round(total_qty, 3), "methods": methods_out}


def create_blend(payload: Dict[str, Any], *, user: str = "system") -> Dict[str, Any]:
    """
    payload = {
      sources: [{lot_id, mt_used}, ...],
      output: {grade, lot_no, qty_mt, methods: {method: {axis: value}}}
    }
    Appends new lot column in Master.xlsx, decrements sources, records blend.
    """
    output = payload.get("output") or {}
    sources = payload.get("sources") or []
    grade = _norm(output.get("grade"))
    lot_no = _norm(output.get("lot_no"))
    qty_out = _to_num(output.get("qty_mt"))
    methods = output.get("methods") or {}
    if not grade or not lot_no:
        raise ValueError("output.grade and output.lot_no are required")
    if not qty_out or qty_out <= 0:
        raise ValueError("output.qty_mt must be > 0")
    if not sources:
        raise ValueError("at least one source lot is required")

    # Sum source mt_used, validate
    parsed = parse_master()
    by_id = {l["lot_id"]: l for l in parsed["lots"]}
    src_resolved: List[Dict[str, Any]] = []
    total_used = 0.0
    for s in sources:
        lid = _norm(s.get("lot_id"))
        mt = _to_num(s.get("mt_used")) or 0.0
        lot = by_id.get(lid)
        if not lot:
            raise ValueError(f"source lot {lid} not found")
        if mt <= 0:
            raise ValueError(f"mt_used for {lid} must be > 0")
        if mt > (lot.get("qty_mt") or 0) + 1e-6:
            raise ValueError(f"mt_used {mt} exceeds available {lot.get('qty_mt')} for {lid}")
        src_resolved.append({"lot": lot, "mt": mt})
        total_used += mt

    wb = _load_wb(read_only=False)
    ws = wb[SHEET_NAME]
    new_col = ws.max_column + 1
    today = datetime.now().strftime("%Y-%m-%d")
    now_iso = datetime.now().isoformat(timespec="seconds")

    # Write new column
    ws.cell(ROW_GRADE, new_col).value = grade
    ws.cell(ROW_LOT, new_col).value = lot_no
    ws.cell(ROW_QTY, new_col).value = round(qty_out, 3)
    ws.cell(ROW_LAST_EDITED, new_col).value = today

    # Method axis values
    for layout_row in parsed["layout"]:
        method = layout_row["method"]
        axis = layout_row["axis"]
        if not method or not axis:
            continue
        v = (methods.get(method) or {}).get(axis)
        nv = _to_num(v)
        if nv is not None:
            ws.cell(layout_row["row"], new_col).value = nv

    # Decrement sources
    audit_new = not AUDIT_FILE.exists()
    edit_dates = _read_edit_dates()
    with AUDIT_FILE.open("a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if audit_new:
            w.writerow(["timestamp", "user", "customer", "lot_id", "lot_no", "col_letter", "prev_qty", "consume_mt", "new_qty"])
        for sr in src_resolved:
            col = sr["lot"]["col"]
            prev = _to_num(ws.cell(ROW_QTY, col).value) or 0.0
            new_qty = round(max(prev - sr["mt"], 0.0), 3)
            ws.cell(ROW_QTY, col).value = new_qty
            ws.cell(ROW_LAST_EDITED, col).value = today
            edit_dates[sr["lot"]["lot_id"]] = today
            w.writerow([now_iso, user, f"BLEND->{lot_no}", sr["lot"]["lot_id"], sr["lot"]["lot_no"], sr["lot"]["col_letter"], prev, sr["mt"], new_qty])

    wb.save(MASTER_FILE)
    _write_edit_dates(edit_dates)

    new_lot_id = f"{lot_no}@{_col_letter(new_col)}"
    record = {
        "blend_id": f"BLEND-{int(datetime.now().timestamp())}",
        "ts": now_iso,
        "user": user,
        "output": {
            "lot_id": new_lot_id,
            "lot_no": lot_no,
            "grade": grade,
            "col_letter": _col_letter(new_col),
            "qty_mt": round(qty_out, 3),
        },
        "sources": [
            {
                "lot_id": sr["lot"]["lot_id"],
                "lot_no": sr["lot"]["lot_no"],
                "grade": sr["lot"]["grade"],
                "mt_used": sr["mt"],
            }
            for sr in src_resolved
        ],
    }

    record["card_url"] = f"/api/v3/master/blend/{record['blend_id']}/card.pdf"

    rec = _read_blend_records()
    rec["blends"].append(record)
    notes = rec.setdefault("notes_by_lot", {})
    for sr in src_resolved:
        lid = sr["lot"]["lot_id"]
        notes.setdefault(lid, []).append({
            "ts": now_iso,
            "mt_used": sr["mt"],
            "into_lot_no": lot_no,
            "into_lot_id": new_lot_id,
            "note": f"{sr['mt']} MT used for blending into {lot_no}",
            "card_url": record["card_url"],
        })
    _write_blend_records(rec)

    try:
        generate_blend_card_pdf(record)
    except Exception:
        pass

    return {"record": record, "total_used_mt": round(total_used, 3)}


if __name__ == "__main__":
    data = parse_master()
    print(f"Methods: {data['methods']}")
    print(f"Lots: {len(data['lots'])}")
    sample = data["lots"][0]
    print(f"Sample lot: {sample['lot_no']} grade={sample['grade']} qty={sample['qty_mt']}")
    print(f"  Present methods: {sample['present_methods']}")
    print(f"  Method I b block: {sample['blocks'].get('Method I b')}")
